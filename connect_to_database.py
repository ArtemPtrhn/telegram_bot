from openpyxl import load_workbook


def insert_sticker(keyword, sticker_id=None, reply_text=None):
    row = stickers_page.max_row + 1
    stickers_page.cell(row=row, column=1).value = keyword
    stickers_page.cell(row=row, column=2).value = sticker_id
    stickers_page.cell(row=row, column=3).value = reply_text
    bd.save('database.xlsx')
    # Добавляем в словари стикеров и ответов
    stickers[keyword] = sticker_id
    replies[keyword] = reply_text


def insert_user(*args):
    '''
    Записывает нового пользователя в БД
    :param args: принимает произвольное количество аргументов
    '''
    user_id = args[0]
    name = args[1]
    sex = args[2]
    grade = args[3]
    row = users_page.max_row + 1
    users_page.cell(row=row, column=1).value = user_id
    users_page.cell(row=row, column=2).value = name
    users_page.cell(row=row, column=3).value = sex
    users_page.cell(row=row, column=4).value = grade

    bd.save('database.xlsx')


def in_database(user: int) -> bool:
    '''
    Возвращает True, если пользователь есть в базе данных database
    '''
    for i in range(2, users_page.max_row + 1):
        if user == users_page.cell(row=row, column=1).value:
            return True
    return False


bd = load_workbook('database.xlsx')
stickers_page = bd['Лист1']
users_page = bd['Users']


stickers = {}
replies = {}

for row in range(2, stickers_page.max_row + 1):
    print(row)
    keyword = stickers_page.cell(row=row, column=1).value
    sticker_id = stickers_page.cell(row=row, column=2).value
    reply_text = stickers_page.cell(row=row, column=3).value
    stickers[keyword] = sticker_id
    replies[keyword] = reply_text

if __name__ == '__main__':
    print(stickers)
    # Тестируем новую функцию
    insert_user(12345, 'artem', 'М', '10')
